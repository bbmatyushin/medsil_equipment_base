from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.contrib import admin


def export_to_excel_formatted(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formatted_data.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Стили для заголовков
    header_font = Font(bold=True, size=12)
    center_aligned = Alignment(horizontal='center')

    # Получаем заголовки из list_display модели админки
    headers = []
    for field_name in modeladmin.list_display:
        if field_name == 'action_checkbox':
            continue  # Пропускаем чекбокс для выбора
        
        # Получаем читаемое название поля
        try:
            field = modeladmin.model._meta.get_field(field_name)
            header = field.verbose_name
        except:
            # Для методов или специальных полей
            if hasattr(modeladmin, field_name):
                attr = getattr(modeladmin, field_name)
                if hasattr(attr, 'short_description'):
                    header = attr.short_description
                else:
                    header = field_name.replace('_', ' ').title()
            else:
                header = field_name.replace('_', ' ').title()
        
        headers.append(str(header))

    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=str(header))
        cell.font = header_font
        cell.alignment = center_aligned

    # Данные
    for row, obj in enumerate(queryset, 2):
        col_num = 1
        for field_name in modeladmin.list_display:
            if field_name == 'action_checkbox':
                continue
            
            try:
                # Получаем значение поля
                if hasattr(obj, field_name):
                    value = getattr(obj, field_name)
                else:
                    value = None
                
                # Если это callable (метод), вызываем его
                if callable(value):
                    value = value()
                
                # Форматируем специальные типы данных
                if hasattr(value, 'strftime'):  # datetime объекты
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                elif value is None:
                    value = ''
                elif isinstance(value, bool):
                    value = 'Да' if value else 'Нет'
                
                ws.cell(row=row, column=col_num, value=str(value))
                
            except Exception as e:
                ws.cell(row=row, column=col_num, value=f"Error: {str(e)}")
            
            col_num += 1

    # Авто-ширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response


export_to_excel_formatted.short_description = "Export to Excel (formatted)"
