from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.contrib import admin
from .models import YourModel


def export_to_excel_formatted(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="formatted_data.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Стили для заголовков
    header_font = Font(bold=True, size=12)
    center_aligned = Alignment(horizontal='center')

    # Заголовки
    headers = ['ID', 'Name', 'Email', 'Created At']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_aligned

    # Данные
    for row, obj in enumerate(queryset, 2):
        ws.cell(row=row, column=1, value=obj.id)
        ws.cell(row=row, column=2, value=obj.name)
        ws.cell(row=row, column=3, value=obj.email)
        ws.cell(row=row, column=4, value=obj.created_at.strftime('%Y-%m-%d %H:%M:%S') if obj.created_at else '')

    # Авто-ширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(response)
    return response


export_to_excel_formatted.short_description = "Export to Excel (formatted)"